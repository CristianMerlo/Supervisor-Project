import os
import glob
import time
import openpyxl
from collections import defaultdict
import motor_supervisor
import seguimiento_ppm
import notificador_telegram

def main():
    procesados_dir = "/home/cristian/Documentos/Supervisor/entrantes/procesados/"
    locales_dir = "/home/cristian/Documentos/Supervisor/Locales/"
    
    pdfs = []
    pdfs.extend(glob.glob(os.path.join(procesados_dir, "*.pdf")))
    for root, dirs, files in os.walk(locales_dir):
        for file in files:
            if file.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, file))
                
    print(f"Total PDFs encontrados: {len(pdfs)}")
    
    # Agrupar por sigla
    por_sigla = defaultdict(list)
    for pdf in pdfs:
        basename = os.path.basename(pdf)
        sigla = None
        partes = basename.split("_")
        if len(partes) >= 3 and partes[0] == "MTZ":
            sigla = partes[1].upper()
        else:
            folder_name = os.path.basename(os.path.dirname(pdf)).upper()
            if len(folder_name) >= 3 and folder_name not in ["PROCESADOS", "LOCALES"]:
                sigla = folder_name
                
        if sigla:
            por_sigla[sigla].append(pdf)
            
    print(f"Siglas únicas encontradas: {len(por_sigla)}")
    
    # 1. Descargar el Excel una sola vez
    print("Descargando archivo Agua Seguimiento.xlsx...")
    if not seguimiento_ppm.descargar_xlsx():
        print("Error descargando el Excel.")
        return
        
    wb = openpyxl.load_workbook(seguimiento_ppm.LOCAL_XLSX_PATH)
    hoja = wb['Franquicias'] if 'Franquicias' in wb.sheetnames else wb.active
    
    cambios = 0
    
    # 2. Procesar el más reciente por sigla
    for i, (sigla, archivos) in enumerate(por_sigla.items()):
        # Ordenar por fecha mtime desc (más reciente primero)
        archivos.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        pdf_reciente = archivos[0]
        
        print(f"[{i+1}/{len(por_sigla)}] Procesando {sigla} -> {os.path.basename(pdf_reciente)}...")
        
        # Extraer con IA (hace delay automático gracias a time.sleep)
        datos, _ = motor_supervisor.parser_hibrido(pdf_reciente)
        
        # Buscar fila en Excel
        fila_encontrada = None
        for row in range(2, hoja.max_row + 1):
            celda_sigla = hoja.cell(row=row, column=2).value
            if celda_sigla and str(celda_sigla).strip().upper() == sigla.upper():
                fila_encontrada = row
                break
                
        if fila_encontrada:
            # Escribir PPM
            ppm = datos.get("ppm", 0)
            if ppm > 0:
                hoja.cell(row=fila_encontrada, column=5).value = ppm
                
            # Escribir X
            filtro = "X" if datos.get("filtro_presente") else ""
            ablandador = "X" if datos.get("ablandador_presente") else ""
            osmosis = "X" if datos.get("osmosis_presente") else ""
            
            hoja.cell(row=fila_encontrada, column=12).value = filtro
            hoja.cell(row=fila_encontrada, column=13).value = ablandador
            hoja.cell(row=fila_encontrada, column=14).value = osmosis
            
            notas = datos.get("observaciones_hidricas", "")
            if notas:
                hoja.cell(row=fila_encontrada, column=16).value = notas
                
            tecnico = datos.get("tecnico", "")
            if tecnico:
                hoja.cell(row=fila_encontrada, column=17).value = tecnico
                
            cambios += 1
            print(f"   -> Actualizado en Excel (Fila {fila_encontrada}).")
        else:
            print(f"   -> [!] Sigla {sigla} no encontrada en Excel.")
            
        # Pausa para no romper el rate limit de la API
        time.sleep(3.5)
        
    # 3. Guardar y subir
    print("Guardando Excel...")
    wb.save(seguimiento_ppm.LOCAL_XLSX_PATH)
    wb.close()
    
    print("Subiendo a Google Drive...")
    if seguimiento_ppm.subir_xlsx():
        print("¡Excel actualizado con éxito!")
        if os.path.exists(seguimiento_ppm.LOCAL_XLSX_PATH):
            os.remove(seguimiento_ppm.LOCAL_XLSX_PATH)
    else:
        print("Error subiendo el Excel.")
        
    msg = f"🛠️ [Antigravity] ¡Backfill Hídrico Completado!\nSe analizaron {len(por_sigla)} locales y se actualizaron {cambios} registros en el Excel (incluyendo Filtros, Ablandadores y PPM) usando el procesamiento por lotes con Gemini."
    notificador_telegram.enviar_alerta(msg)
    print("Notificación enviada.")

if __name__ == "__main__":
    main()

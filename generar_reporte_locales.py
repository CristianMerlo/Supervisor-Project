import os
import re
import sys
import time
import sqlite3
import smtplib
import subprocess
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).parent
SCOPES = ['https://www.googleapis.com/auth/drive']
MOSTAZA_LOCALES_FOLDER_ID = "1iOGWgu04vtGRv2QBpmxhT5b5NROkJHR8"

def cargar_env():
    ruta_env = BASE_DIR / ".env"
    if ruta_env.exists():
        with open(ruta_env, "r", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                if not linea or linea.startswith("#") or "=" not in linea:
                    continue
                k, v = linea.split("=", 1)
                os.environ[k.strip()] = v.strip()

def obtener_servicio_drive():
    ruta_credenciales = BASE_DIR / "credentials.json"
    if not ruta_credenciales.exists():
        raise FileNotFoundError(f"No se encontró credentials.json en: {ruta_credenciales}")
    creds = Credentials.from_service_account_file(str(ruta_credenciales), scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)

def obtener_locales_db():
    db_path = Path("/home/cristian/Documentos/Supervisor/supervisor_local.db")
    locales_map = {}
    if db_path.exists():
        try:
            conn = sqlite3.connect(str(db_path))
            cursor = conn.cursor()
            cursor.execute("SELECT sigla, nombre FROM locales")
            for row in cursor.fetchall():
                sigla = row[0].strip().upper()
                nombre = row[1].strip()
                locales_map[sigla] = nombre
            conn.close()
        except Exception as e:
            print(f"[DB-ERROR] No se pudo leer SQLite: {e}")
    return locales_map

def contar_pdfs_en_carpeta(service, folder_id):
    # 1. Buscar si hay subcarpeta 'Reportes'
    query_sub = f"mimeType='application/vnd.google-apps.folder' and '{folder_id}' in parents and name = 'Reportes' and trashed = false"
    try:
        res_sub = service.files().list(q=query_sub, fields='files(id)').execute()
        subs = res_sub.get('files', [])
    except Exception as e:
        print(f"[DRIVE-ERROR] Error buscando subcarpeta 'Reportes' para {folder_id}: {e}")
        subs = []
        
    parent_ids = [folder_id]
    if subs:
        parent_ids.append(subs[0]['id'])
        
    total_pdfs = 0
    # 2. Contar archivos PDF en las carpetas
    for pid in parent_ids:
        query_files = f"mimeType='application/pdf' and '{pid}' in parents and trashed = false"
        page_token = None
        while True:
            try:
                res_files = service.files().list(
                    q=query_files,
                    fields='nextPageToken, files(id)',
                    pageSize=100,
                    pageToken=page_token
                ).execute()
                total_pdfs += len(res_files.get('files', []))
                page_token = res_files.get('nextPageToken')
                if not page_token:
                    break
            except Exception as e:
                print(f"[DRIVE-ERROR] Error listando PDFs en {pid}: {e}")
                break
    return total_pdfs

def generar_excel(datos, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Informes"
    
    # Grid lines visibles

    
    # Encabezados
    headers = ["Sigla", "Nombre Local", "Informes en Drive", "Estado de Información"]
    ws.append(headers)
    
    # Datos
    for d in datos:
        ws.append([d['sigla'], d['nombre'], d['cantidad'], d['estado']])
        
    # Estilizado
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_data = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Formatear Cabecera
    for col_idx in range(1, 5):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 3, 4] else align_left
        
    # Formatear Datos
    for row_idx in range(2, len(datos) + 2):
        ws.cell(row=row_idx, column=1).alignment = align_center # Sigla
        ws.cell(row=row_idx, column=2).alignment = align_left   # Nombre
        ws.cell(row=row_idx, column=3).alignment = align_right  # Cantidad
        ws.cell(row=row_idx, column=4).alignment = align_center # Estado
        
        # Bordes
        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).border = border_data
            
    # Fila de Totales
    total_row_idx = len(datos) + 2
    ws.cell(row=total_row_idx, column=1, value="TOTALES").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=total_row_idx, column=3, value=f"=SUM(C2:C{total_row_idx-1})").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=total_row_idx, column=3).alignment = align_right
    
    double_bottom = Border(top=thin_side, bottom=Side(border_style="double", color="000000"))
    for col_idx in range(1, 5):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.border = double_bottom
        
    # Ancho de columnas automático
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_path)

def generar_url_grafico(top_locales):
    # Generar URL de Google Charts (gráfico de barras vertical / de columna)
    if not top_locales:
        return ""
    
    # Invertir para mostrar de mayor a menor en gráfico de barras horizontales (bhs)
    top_locales_sorted = list(reversed(top_locales))
    labels = "|".join([x['sigla'] for x in top_locales_sorted])
    values = ",".join([str(x['cantidad']) for x in top_locales_sorted])
    max_val = max([x['cantidad'] for x in top_locales])
    
    url = (
        f"https://chart.googleapis.com/chart?"
        f"cht=bhs&" # Horizontal bar chart
        f"chs=600x300&"
        f"chco=4F81BD&" # Azul acero
        f"chd=t:{values}&"
        f"chds=0,{max_val+2}&"
        f"chxt=y,x&"
        f"chxl=0:|{labels}&"
        f"chtt=Top+10+Locales+con+mas+Informes&"
        f"chg=20,0&"
        f"chf=bg,s,FFFFFF"
    )
    return url

def enviar_correo_html(asunto, cuerpo_html, adjunto_path):
    gmail_user = os.getenv("GMAIL_USER")
    gmail_pass = os.getenv("GMAIL_APP_PASSWORD")
    destinatario = os.getenv("CORREO_CORPORATIVO", gmail_user)
    
    if not gmail_user or not gmail_pass:
        print("[MAIL] Error: Credenciales de Gmail no configuradas.")
        return False
        
    try:
        if "[Antigravity]" not in asunto and "[Supervisor]" not in asunto:
            asunto = f"[Antigravity] {asunto}"
            
        msg = MIMEMultipart()
        msg['From'] = gmail_user
        msg['To'] = destinatario
        msg['Subject'] = asunto
        
        msg.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
        
        # Adjuntar archivo Excel
        with open(adjunto_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={Path(adjunto_path).name}"
        )
        msg.attach(part)
        
        server = smtplib.SMTP('smtp.gmail.com', 587, timeout=15)
        server.starttls()
        server.login(gmail_user, gmail_pass)
        server.sendmail(gmail_user, destinatario, msg.as_string())
        server.quit()
        print(f"[MAIL] Reporte enviado por mail a {destinatario}.")
        return True
    except Exception as e:
        print(f"[MAIL-ERROR] Falló envío de mail: {e}")
        return False

def enviar_telegram(mensaje):
    try:
        subprocess.run(
            ["/home/cristian/Documentos/Supervisor/notify_telegram.sh", mensaje],
            check=True
        )
        print("[TELEGRAM] Resumen enviado con éxito.")
    except Exception as e:
        print(f"[TG-ERROR] Falló notificación Telegram: {e}")

def main():
    print("=== PROCESANDO REPORTE DE INFORMES EN GOOGLE DRIVE ===")
    cargar_env()
    
    try:
        service = obtener_servicio_drive()
    except Exception as e:
        print(f"[CRITICAL] Error conectando a Google Drive: {e}")
        return
        
    locales_db = obtener_locales_db()
    print(f"Cargados {len(locales_db)} locales oficiales de la base de datos local.")
    
    # 1. Recuperar todas las carpetas del local en la carpeta raíz
    carpetas_drive = []
    page_token = None
    query_folders = f"mimeType='application/vnd.google-apps.folder' and '{MOSTAZA_LOCALES_FOLDER_ID}' in parents and trashed = false"
    
    while True:
        try:
            res = service.files().list(
                q=query_folders,
                fields='nextPageToken, files(id, name)',
                pageSize=100,
                pageToken=page_token
            ).execute()
            carpetas_drive.extend(res.get('files', []))
            page_token = res.get('nextPageToken')
            if not page_token:
                break
        except Exception as e:
            print(f"[CRITICAL] Error listando carpetas en Drive: {e}")
            return
            
    print(f"Detectadas {len(carpetas_drive)} carpetas de sucursales en Drive.")
    
    # 2. Contar informes por cada carpeta
    informes_por_local = {}
    siglas_procesadas = set()
    
    for idx, folder in enumerate(carpetas_drive):
        folder_name = folder['name']
        folder_id = folder['id']
        
        # Ignorar la carpeta de la bandeja de entrada
        if folder_name == "001_Bandeja_de_Entrada":
            continue
            
        # Extraer sigla
        m = re.match(r"^\[(.*?)\]", folder_name)
        if m:
            sigla = m.group(1).strip().upper()
        else:
            # Fallback si no tiene corchetes
            sigla = folder_name.split("-")[0].strip().upper()
            
        siglas_procesadas.add(sigla)
        nombre_oficial = locales_db.get(sigla, folder_name.split("-")[-1].strip())
        
        # Contar PDFs
        cant_pdfs = contar_pdfs_en_carpeta(service, folder_id)
        
        informes_por_local[sigla] = {
            'sigla': sigla,
            'nombre': nombre_oficial,
            'cantidad': cant_pdfs,
            'estado': "Activo" if cant_pdfs > 0 else "Sin Informes"
        }
        print(f"[{idx+1}/{len(carpetas_drive)}] Local: {sigla} | PDFs: {cant_pdfs}")
        
    # 3. Identificar locales de la DB que no tienen carpeta o tienen 0 reportes
    for sigla, nombre in locales_db.items():
        if sigla not in informes_por_local:
            informes_por_local[sigla] = {
                'sigla': sigla,
                'nombre': nombre,
                'cantidad': 0,
                'estado': "Sin Carpeta/Informes"
            }
            
    # Convertir a lista y ordenar de mayor a menor
    lista_reportes = list(informes_por_local.values())
    lista_reportes.sort(key=lambda x: x['cantidad'], reverse=True)
    
    # Calcular totales
    total_informes = sum(x['cantidad'] for x in lista_reportes)
    locales_con_informes = sum(1 for x in lista_reportes if x['cantidad'] > 0)
    locales_sin_informes = sum(1 for x in lista_reportes if x['cantidad'] == 0)
    
    print(f"\nResumen: Total informes: {total_informes} | Con Informes: {locales_con_informes} | Sin Informes: {locales_sin_informes}")
    
    # 4. Generar planilla Excel
    excel_path = "/tmp/Reporte_Informes_Drive.xlsx"
    generar_excel(lista_reportes, excel_path)
    
    # 5. Top 10 locales con más reportes para gráfico
    top_locales = [x for x in lista_reportes if x['cantidad'] > 0][:10]
    chart_url = generar_url_grafico(top_locales)
    
    # 6. Construir cuerpo del correo HTML
    tabla_html_rows = ""
    for idx, d in enumerate(lista_reportes):
        # Limitar la tabla HTML a los primeros 25 y los sin informes para que no sea gigante
        if idx < 15 or d['cantidad'] == 0:
            bg_color = "#f9f9f9" if idx % 2 == 0 else "#ffffff"
            text_color = "#d9534f" if d['cantidad'] == 0 else "#333333"
            tabla_html_rows += f"""
            <tr style="background-color: {bg_color}; color: {text_color};">
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-weight: bold;">{d['sigla']}</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{d['nombre']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: right; font-weight: bold;">{d['cantidad']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{d['estado']}</td>
            </tr>
            """
        elif idx == 15:
            tabla_html_rows += """
            <tr style="background-color: #f1f1f1; text-align: center;">
                <td colspan="4" style="padding: 10px; font-style: italic; color: #777;">... (ver el resto de los locales en la planilla Excel adjunta) ...</td>
            </tr>
            """
            
    cuerpo_html = f"""
    <html>
    <head>
        <meta charset="utf-8">
    </head>
    <body style="font-family: Arial, sans-serif; color: #333; line-height: 1.6; max-width: 700px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #1F497D; border-bottom: 2px solid #1F497D; padding-bottom: 10px; margin-top: 0;">
            📊 Reporte de Informes Técnicos por Sucursal
        </h2>
        <p>Cristian, a continuación se presenta el análisis de los informes PDF de mantenimiento cargados en <strong>Google Drive</strong> (carpeta raíz <em>"Mostaza Locales"</em>).</p>
        
        <!-- Tarjetas de Resumen -->
        <table style="width: 100%; margin: 20px 0; border-collapse: collapse; text-align: center;">
            <tr>
                <td style="width: 33%; padding: 15px; background-color: #E2EFDA; border: 1px solid #C6E0B4; border-radius: 4px;">
                    <div style="font-size: 24px; font-weight: bold; color: #375623;">{total_informes}</div>
                    <div style="font-size: 12px; color: #548235; font-weight: bold;">Total Informes</div>
                </td>
                <td style="width: 33%; padding: 15px; background-color: #DDEBF7; border: 1px solid #B4C6E7; border-radius: 4px;">
                    <div style="font-size: 24px; font-weight: bold; color: #1F4E78;">{locales_con_informes}</div>
                    <div style="font-size: 12px; color: #2F5597; font-weight: bold;">Locales con Informes</div>
                </td>
                <td style="width: 33%; padding: 15px; background-color: #FCE4D6; border: 1px solid #F8CBAD; border-radius: 4px;">
                    <div style="font-size: 24px; font-weight: bold; color: #C65911;">{locales_sin_informes}</div>
                    <div style="font-size: 12px; color: #833C0C; font-weight: bold;">Locales sin Informes</div>
                </td>
            </tr>
        </table>

        <!-- Gráfico Estadístico -->
        {"<div style='text-align: center; margin: 25px 0;'><img src='" + chart_url + "' alt='Top 10 Locales' style='max-width:100%; height:auto; border: 1px solid #ddd; padding: 5px; border-radius: 4px;' /></div>" if chart_url else ""}

        <h3 style="color: #1F497D; margin-top: 25px;">📋 Detalle de Sucursales (Primeros 15 y Locales sin Informes)</h3>
        <table style="width: 100%; border-collapse: collapse; font-size: 13px;">
            <thead>
                <tr style="background-color: #1F497D; color: white;">
                    <th style="padding: 10px; border: 1px solid #ddd; text-align: center;">Sigla</th>
                    <th style="padding: 10px; border: 1px solid #ddd; text-align: left;">Nombre Local</th>
                    <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">Informes</th>
                    <th style="padding: 10px; border: 1px solid #ddd; text-align: center;">Estado</th>
                </tr>
            </thead>
            <tbody>
                {tabla_html_rows}
            </tbody>
        </table>
        
        <p style="margin-top: 20px; font-size: 12px; color: #777;">
            * Se adjunta la planilla de Excel con el 100% de los locales analizados y el detalle consolidado.
        </p>
    </body>
    </html>
    """
    
    # 7. Enviar Mail
    asunto = "Reporte de Informes Técnicos por Sucursal en Drive"
    enviar_correo_html(asunto, cuerpo_html, excel_path)
    
    # 8. Construir mensaje corto para Telegram
    # Top 5 con informes y lista corta de sin informes
    top_5_str = "\n".join([f"• *{x['sigla']}*: {x['cantidad']} informes" for x in lista_reportes[:5]])
    sin_inf_siglas = [x['sigla'] for x in lista_reportes if x['cantidad'] == 0]
    sin_inf_str = ", ".join(sin_inf_siglas[:10])
    if len(sin_inf_siglas) > 10:
        sin_inf_str += f" y {len(sin_inf_siglas)-10} más"
        
    mensaje_telegram = (
        f"🧠 *[Hermes] Resumen de Informes en Google Drive*\n\n"
        f"• *Total de informes*: {total_informes}\n"
        f"• *Locales con informes*: {locales_con_informes}\n"
        f"• *Locales sin informes*: {locales_sin_informes}\n\n"
        f"🔝 *Top 5 locales con más informes*:\n{top_5_str}\n\n"
        f"⚠️ *Locales sin informes*:\n{sin_inf_str if sin_inf_siglas else 'Ninguno'}\n\n"
        f"✉️ _Se envió el reporte completo en formato HTML y la planilla de Excel adjunta a tu correo corporativo._"
    )
    enviar_telegram(mensaje_telegram)
    
    # Limpieza temporal
    try:
        os.remove(excel_path)
    except Exception:
        pass

if __name__ == "__main__":
    main()

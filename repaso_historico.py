import os
import glob
import subprocess
import openpyxl
from dotenv import load_dotenv
from collections import defaultdict

load_dotenv("/home/cristian/PROYECTOS/Supervisor-Project/.env")
from clasificar_bandeja_entrada import extraer_ppm_con_gemini
import seguimiento_ppm

def main():
    procesados_dir = "/home/cristian/Documentos/Supervisor/entrantes/procesados/"
    locales_dir = "/home/cristian/Documentos/Supervisor/Locales/"
    
    # Recolectar todos los PDFs
    pdfs = []
    pdfs.extend(glob.glob(os.path.join(procesados_dir, "*.pdf")))
    # En Locales puede haber subcarpetas
    for root, dirs, files in os.walk(locales_dir):
        for file in files:
            if file.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, file))
                
    print(f"Total PDFs encontrados: {len(pdfs)}")
    
    # Agrupar por sigla
    por_sigla = defaultdict(list)
    for pdf in pdfs:
        basename = os.path.basename(pdf)
        # Formato esperado: MTZ_SIGLA_YYYY-MM-DD.pdf o estar en carpeta SIGLA
        sigla = None
        partes = basename.split("_")
        if len(partes) >= 3 and partes[0] == "MTZ":
            sigla = partes[1].upper()
        else:
            # Intentar sacar de la carpeta
            folder_name = os.path.basename(os.path.dirname(pdf)).upper()
            if len(folder_name) >= 3 and folder_name != "PROCESADOS" and folder_name != "LOCALES":
                sigla = folder_name
                
        if sigla:
            por_sigla[sigla].append(pdf)
            
    print(f"Siglas únicas encontradas: {len(por_sigla)}")
    
    updates = {}
    
    # Extraer el PPM más reciente para cada sigla
    for sigla, archivos in por_sigla.items():
        # Ordenar por fecha de modificación descendente
        archivos.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        
        for pdf in archivos:
            try:
                texto = subprocess.check_output(['pdftotext', pdf, '-']).decode('utf-8', errors='ignore')
            except Exception as e:
                continue
                
            import re
            ppm = None
            match = re.search(r'PPM:\s*(\d+(\.\d+)?)', texto, re.IGNORECASE)
            if match:
                ppm = float(match.group(1))
            
            if ppm is not None:
                print(f"[+] Sigla {sigla}: PPM {ppm} encontrado en {os.path.basename(pdf)}")
                updates[sigla] = ppm
                break # Ya encontramos el más reciente, pasamos a la siguiente sigla
                
    if not updates:
        print("No se encontraron actualizaciones pendientes.")
        return
        
    print(f"\nSe encontraron {len(updates)} mediciones de PPM para actualizar.")
    print("Iniciando volcado masivo en Excel...")
    
    # Descargar Excel
    if not seguimiento_ppm.descargar_xlsx():
        print("Error al descargar Excel")
        return
        
    wb = openpyxl.load_workbook(seguimiento_ppm.LOCAL_XLSX_PATH)
    hoja = wb['Franquicias'] if 'Franquicias' in wb.sheetnames else wb.active
    
    cambios_realizados = 0
    
    for row in range(2, hoja.max_row + 1):
        celda_sigla = hoja.cell(row=row, column=2).value
        if celda_sigla:
            s = str(celda_sigla).strip().upper()
            if s in updates:
                nuevo_ppm = float(updates[s])
                ppm_anterior = hoja.cell(row=row, column=5).value
                
                # Reglas de alerta en modo masivo: solo logueamos, pero actualizamos igual 
                # (o podriamos no actualizar si cae a 0, para ser super seguros)
                if nuevo_ppm == 0:
                    print(f"  [ALERTA] Local {s} bajó a 0. Se registra igual por instrucción.")
                elif ppm_anterior and float(ppm_anterior) > 0 and (nuevo_ppm < float(ppm_anterior) * 0.5):
                    print(f"  [ALERTA] Local {s} cayó >50% (de {ppm_anterior} a {nuevo_ppm}). Se registra igual.")
                    
                hoja.cell(row=row, column=5).value = nuevo_ppm
                cambios_realizados += 1
                
    wb.save(seguimiento_ppm.LOCAL_XLSX_PATH)
    wb.close()
    
    print(f"Volcado en Excel completado. {cambios_realizados} filas modificadas.")
    
    # Subir Excel
    if seguimiento_ppm.subir_xlsx():
        print("Excel subido correctamente a Google Drive.")
        if os.path.exists(seguimiento_ppm.LOCAL_XLSX_PATH):
            os.remove(seguimiento_ppm.LOCAL_XLSX_PATH)
    else:
        print("Error al subir el archivo.")

if __name__ == "__main__":
    main()

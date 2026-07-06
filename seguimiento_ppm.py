import os
import io
import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
import notificador_telegram

# Configuración
FILE_ID = '1qhj1abI3w_EUCgRIQ4Nd4v6d8Ng__4na'
CREDENTIALS_FILE = '/home/cristian/Documentos/Supervisor/credentials.json'
LOCAL_XLSX_PATH = '/home/cristian/PROYECTOS/Supervisor-Project/brain/AGUA_SEGUIMIENTO_temp.xlsx'
SCOPES = ['https://www.googleapis.com/auth/drive']

def get_drive_service():
    creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)

def descargar_xlsx():
    service = get_drive_service()
    request = service.files().get_media(fileId=FILE_ID)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        
    fh.seek(0)
    with open(LOCAL_XLSX_PATH, 'wb') as f:
        f.write(fh.read())
    return True

def subir_xlsx():
    service = get_drive_service()
    media = MediaFileUpload(LOCAL_XLSX_PATH, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    service.files().update(fileId=FILE_ID, media_body=media).execute()
    return True

def procesar_medicion_ppm(sigla, nuevo_ppm):
    """
    Busca la sigla en la solapa 'Franquicias' y actualiza la columna E (PPM)
    si pasa las reglas de validación.
    """
    # 1. Descargar el archivo
    if not descargar_xlsx():
        return False, "Error al descargar el archivo de Drive."
        
    # 2. Abrir con openpyxl
    try:
        wb = openpyxl.load_workbook(LOCAL_XLSX_PATH)
        if 'Franquicias' not in wb.sheetnames:
            # Quizás esté en minúsculas
            hoja = wb.active
            for s in wb.sheetnames:
                if s.lower() == 'franquicias':
                    hoja = wb[s]
                    break
        else:
            hoja = wb['Franquicias']
            
        # 3. Buscar la fila por la sigla (Columna B)
        fila_encontrada = None
        ppm_anterior = None
        
        # Asumiendo que la Columna B es la 2 y la E es la 5.
        for row in range(2, hoja.max_row + 1):
            celda_sigla = hoja.cell(row=row, column=2).value
            if celda_sigla and str(celda_sigla).strip().upper() == sigla.upper():
                fila_encontrada = row
                ppm_anterior = hoja.cell(row=row, column=5).value
                break
                
        if not fila_encontrada:
            return False, f"No se encontró el local con sigla {sigla} en la planilla."
            
        # Convertir a números para comparar
        try:
            nuevo_ppm_val = float(nuevo_ppm)
        except ValueError:
            return False, f"El nuevo valor de PPM '{nuevo_ppm}' no es un número válido."
            
        ppm_ant_val = 0.0
        try:
            if ppm_anterior is not None:
                ppm_ant_val = float(ppm_anterior)
        except ValueError:
            pass
            
        # 4. Validar umbral (caída a 0 o caída brusca > 50%)
        necesita_aprobacion = False
        razon_alerta = ""
        
        if nuevo_ppm_val == 0:
            necesita_aprobacion = True
            razon_alerta = "El valor nuevo es 0"
        elif ppm_ant_val > 0 and (nuevo_ppm_val < ppm_ant_val * 0.5):
            necesita_aprobacion = True
            razon_alerta = f"Caída abrupta (>50%). Anterior: {ppm_ant_val}, Nuevo: {nuevo_ppm_val}"
            
        if necesita_aprobacion:
            mensaje = f"⚠️ [PPM] {razon_alerta} para el local {sigla}. ¿Deseas aprobar la actualización en la planilla?"
            decision = notificador_telegram.solicitar_aprobacion(mensaje, timeout=3600)
            if decision != "approved":
                return False, "Actualización rechazada por el usuario o tiempo agotado."
                
        # 5. Escribir el nuevo valor
        hoja.cell(row=fila_encontrada, column=5).value = nuevo_ppm_val
        wb.save(LOCAL_XLSX_PATH)
        wb.close()
        
        # 6. Subir a Google Drive
        subir_xlsx()
        
        # Limpiar temporal
        if os.path.exists(LOCAL_XLSX_PATH):
            os.remove(LOCAL_XLSX_PATH)
            
        return True, f"Medición de PPM ({nuevo_ppm}) actualizada exitosamente para {sigla}."
        
    except Exception as e:
        return False, f"Error al procesar el Excel: {e}"

def actualizar_datos_hidricos(sigla, datos_extraidos):
    """
    Busca la sigla en 'Franquicias' y actualiza L, M, N (X), P (Notas) y Q (Técnico).
    """
    if not descargar_xlsx():
        return False, "Error al descargar el archivo de Drive."
        
    try:
        wb = openpyxl.load_workbook(LOCAL_XLSX_PATH)
        if 'Franquicias' not in wb.sheetnames:
            hoja = wb.active
            for s in wb.sheetnames:
                if s.lower() == 'franquicias':
                    hoja = wb[s]
                    break
        else:
            hoja = wb['Franquicias']
            
        fila_encontrada = None
        for row in range(2, hoja.max_row + 1):
            celda_sigla = hoja.cell(row=row, column=2).value
            if celda_sigla and str(celda_sigla).strip().upper() == sigla.upper():
                fila_encontrada = row
                break
                
        if not fila_encontrada:
            return False, f"No se encontró el local con sigla {sigla} en la planilla."
            
        ppm = datos_extraidos.get("ppm", 0)
        if ppm > 0:
            hoja.cell(row=fila_encontrada, column=5).value = ppm
            
        filtro = "X" if datos_extraidos.get("filtro_presente") else ""
        ablandador = "X" if datos_extraidos.get("ablandador_presente") else ""
        osmosis = "X" if datos_extraidos.get("osmosis_presente") else ""
        
        hoja.cell(row=fila_encontrada, column=12).value = filtro
        hoja.cell(row=fila_encontrada, column=13).value = ablandador
        hoja.cell(row=fila_encontrada, column=14).value = osmosis
        
        notas = datos_extraidos.get("observaciones_hidricas", "")
        if notas:
            hoja.cell(row=fila_encontrada, column=16).value = notas
            
        tecnico = datos_extraidos.get("tecnico", "")
        if tecnico:
            hoja.cell(row=fila_encontrada, column=17).value = tecnico
            
        wb.save(LOCAL_XLSX_PATH)
        wb.close()
        
        subir_xlsx()
        
        import os
        if os.path.exists(LOCAL_XLSX_PATH):
            os.remove(LOCAL_XLSX_PATH)
            
        return True, f"Datos hídricos de {sigla} actualizados."
    except Exception as e:
        return False, f"Error actualizando Excel hídrico: {e}"

def adjuntar_evidencia_visual(sigla, diagnostico, url_foto):
    """
    Agrega el diagnóstico y la URL de la foto en la columna P (Estado Reportado)
    para el local especificado por la sigla.
    """
    if not descargar_xlsx():
        return False, "Error descargando el Excel."
        
    try:
        wb = openpyxl.load_workbook(LOCAL_XLSX_PATH)
        hoja = wb['Franquicias'] if 'Franquicias' in wb.sheetnames else wb.active
        
        fila_encontrada = None
        for row in range(2, hoja.max_row + 1):
            celda_sigla = hoja.cell(row=row, column=2).value
            if celda_sigla and str(celda_sigla).strip().upper() == sigla.upper():
                fila_encontrada = row
                break
                
        if not fila_encontrada:
            return False, f"Local {sigla} no encontrado en la planilla."
            
        # Obtener comentario actual de la Columna 16 (P)
        comentario_actual = hoja.cell(row=fila_encontrada, column=16).value or ""
        
        # Armar la nueva evidencia
        nueva_evidencia = f"[EVIDENCIA VISUAL] {diagnostico} | Ver foto: {url_foto}"
        
        if comentario_actual:
            hoja.cell(row=fila_encontrada, column=16).value = f"{comentario_actual}\n{nueva_evidencia}"
        else:
            hoja.cell(row=fila_encontrada, column=16).value = nueva_evidencia
            
        wb.save(LOCAL_XLSX_PATH)
        wb.close()
        subir_xlsx()
        
        if os.path.exists(LOCAL_XLSX_PATH):
            os.remove(LOCAL_XLSX_PATH)
            
        return True, "Evidencia adjuntada exitosamente."
    except Exception as e:
        return False, f"Error actualizando Excel: {e}"

if __name__ == "__main__":
    # Prueba rápida
    import sys
    if len(sys.argv) > 2:
        res, msg = procesar_medicion_ppm(sys.argv[1], sys.argv[2])
        print(msg)
    else:
        print("Uso: python3 seguimiento_ppm.py <SIGLA> <NUEVO_PPM>")
